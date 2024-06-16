import os
import re
import sys
import requests
import pandas as pd
import time
from DB_Connection import DB_Connection
import json
import sqlite3
import logging
import numpy as np
import psutil
import json
from openpyxl import load_workbook
from openpyxl.styles import numbers

class Filling_Links:
 
    def __init__(self, company_CIKs,folder_path,user_agent):
            self.company_CIKs = company_CIKs
            self.folder_path = folder_path
            self.user_agent = user_agent
            self.logger = logging.getLogger(__name__)
         

    def retrieve_companyfacts_json(self,cik_number):
        """
        Retrieves company facts data in JSON format for the specified CIK number from the SEC API.

        Args:
            cik_number (str): The CIK (Central Index Key) number of the company.

        Returns:
            dict or None: The company facts data in JSON format if retrieval is successful, 
                        None if there was an error.

        Example:
            json_data = retrieve_companyfacts_json("123456789")
            if json_data:
                print("Company facts data retrieved successfully")
                # Process the JSON data
            else:
                print("Failed to retrieve company facts data")
        """
        time.sleep(0.1)
        api_url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik_number}.json"
        print(api_url)
        try:
            response = requests.get(api_url, headers={"User-Agent":self.user_agent})
            response.raise_for_status()  # Raises an exception for non-200 status codes
            json_data = json.loads(response.content)
            return json_data
        except requests.exceptions.RequestException as e:
            print(f"Warning: Error occurred while retrieving company facts for CIK: {cik_number}")
            print(f"Error details: {str(e)}")
            return None
        except json.JSONDecodeError as e:
            print(f"Warning: Error occurred while parsing JSON response for CIK: {cik_number}")
            print(f"Error details: {str(e)}")
            return None

    def is_file_open(self,file_path):
        """
        Checks if a file specified by `file_path` is currently open by any process.
        Args:
            file_path (str): The path of the file to be checked.

        Returns:
            bool: True if the file is open, False otherwise.

        Example:
            is_open = is_file_open("path/to/file.txt")
            if is_open:
                print("File is open")
            else:
                print("File is not open")
        """
        if not os.path.isfile(file_path):
            print(f"Invalid file path: {file_path}")
            return False
        try:
            for proc in psutil.process_iter(['name', 'pid', 'open_files']):
                for file_info in proc.info['open_files']:
                    if file_info.path == file_path:
                            return True
            return False
        except Exception as e:
            print(f"An error occurred while checking if file is open: {str(e)}")
            return False

    def sanitize_filename(self, filename):
        """
        Sanitizes the given filename by removing special characters and invalid characters.

        Args:
            filename (str): The original filename to be sanitized.

        Returns:
            str: The sanitized filename without special characters.

        Example:
            sanitized = sanitize_filename("file<>name.txt")
            print(sanitized)
            # Output: "filename.txt"
        """
        sanitized_filename = re.sub(r'[<>:"/\\|?*]', '', filename)
        return sanitized_filename
    
    def format_values_as_usd(self, conn):
        """
            Formats values in the 'val' column of tables containing 'usd' in their name to USD currency format.

            Args:
                conn: The SQLite database connection object.

            Returns:
                None

            Example:
                import sqlite3

                # Create a connection to the SQLite database
                conn = sqlite3.connect("mydatabase.db")

                # Format values as USD
                format_values_as_usd(conn)

                # Close the database connection
                conn.close()
            """       
        c = conn.cursor()
        c.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = c.fetchall()

        for table in tables:
            table_name = table[0]
            if "usd" in table_name.lower():
                c.execute(f"PRAGMA table_info({table_name});")
                columns = c.fetchall()
                headers = [column[1] for column in columns]

                if "val" in headers:
                    c.execute(f"SELECT * FROM {table_name};")
                    rows = c.fetchall()
                    rows = [list(row) for row in rows]  # Convert rows to lists

                    for row in rows:
                        if "val" in headers:
                            val_index = headers.index("val")
                            val_value = row[val_index]
                            if isinstance(val_value, (int, float)):
                                row[val_index] = f"${val_value:,.2f}"

                    c.execute(f"DELETE FROM {table_name};")
                    c.executemany(f"INSERT INTO {table_name} VALUES ({','.join(['?'] * len(headers))});", rows)

        conn.commit()

    def get_companyfacts_json_db(self):
        """
        Retrieves company facts JSON data for each company CIK number and stores it in a SQLite database.
        
        Example:
            # Create an instance of the class
            obj = MyClass()
            
            # Retrieve company facts JSON data and store it in a database
            obj.get_companyfacts_json_db()
        """
        
        try:
            for Company_CIK_Number in self.company_CIKs:
                json_data = self.retrieve_companyfacts_json(Company_CIK_Number)
                if json_data is None:
                    continue
                
                Primary_Name = json_data['entityName']
                Primary_Name = re.sub(r'[<>:"/\\|?*]', '', Primary_Name)
                db_path = os.path.join(self.folder_path, f"{Primary_Name}.db")
                
                connection_ToFolder = DB_Connection(Primary_Name, self.folder_path)
                connection_ToFolder.create_folder()
                
                self.Get_SubmissionContent_Json(db_path, Company_CIK_Number)
                
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute("DROP TABLE IF EXISTS AvailableDataToExtract")
                c.execute("DROP TABLE IF EXISTS table_of_contents")
                
                tree_structure = None
                for tree_name in ['us-gaap', 'ifrs-full']:
                    if tree_name in json_data['facts']:
                        tree_structure = tree_name
                        break
                
                if tree_structure is None:
                    print(f"Warning: Unknown JSON tree structure for CIK: {Company_CIK_Number}")
                    continue
                
                c.execute('''
                    CREATE TABLE IF NOT EXISTS Metadata_Table (
                        table_name TEXT PRIMARY KEY,
                        array_parameter1 TEXT,
                        label_name TEXT,
                        description_name TEXT,
                        unit_name Text
                    );
                ''')
                
                for item in json_data['facts'][tree_structure]:
                    array_parameter1 = item
                    label_name = json_data['facts'][tree_structure][array_parameter1]["label"]
                    description_name = json_data['facts'][tree_structure][array_parameter1]['description']
                    
                    for unit_item in json_data['facts'][tree_structure][array_parameter1]['units']:
                        unit_name = unit_item
                        table_name = unit_item
                        table_name_unique = f'{array_parameter1} / Unit {table_name}'
                        table_name_unique = re.sub(r'\W+', '_', table_name_unique)
                        
                        c.execute(f"DROP TABLE IF EXISTS {table_name_unique}")
                        df1 = pd.DataFrame(json_data['facts'][tree_structure][array_parameter1]['units'][unit_item])
                        df1.to_sql(table_name_unique, conn, if_exists="replace")
                        
                        # Insert metadata into Metadata_Table
                        metadata_query = '''
                            INSERT OR REPLACE INTO Metadata_Table (table_name, array_parameter1, label_name, description_name, unit_name)
                            VALUES (?, ?, ?, ?, ?)
                        '''
                        c.execute(metadata_query, (table_name_unique, array_parameter1, label_name, description_name,unit_name))
                
                c.execute('''
                    CREATE TABLE IF NOT EXISTS table_of_contents (
                        id INTEGER PRIMARY KEY,
                        table_name TEXT
                    );
                ''')
                
                c.execute('''
                    INSERT INTO table_of_contents (table_name) 
                    SELECT name FROM sqlite_master WHERE type='table';
                ''')
                
                conn.commit()
                # Format values as USD
                self.format_values_as_usd(conn)
                conn.close()
            
        except Exception as e:
            print(f"Error: {e}")
            sys.exit(1)

    def format_values_as_usd_excel(self, file_path):
        """
            Formats the values in the "Val" column of each worksheet in the Excel file as USD currency.

            Args:
                file_path (str): The path to the Excel file.

            Example:
                # Create an instance of the class
                obj = MyClass()

                # Format values as USD currency in the Excel file
                obj.format_values_as_usd_excel('example.xlsx')
        """
        # Load the workbook
        workbook = load_workbook(filename=file_path)

        for worksheet in workbook.worksheets:
            # Check cell B4 for unit "USD"
            cell_value = worksheet['B4'].value
            if cell_value and cell_value.lower() == "usd":
                start_row = 8
                # Find the "Val" header column
                val_column_index = None
                header_row = worksheet[8]

                for col_index, cell in enumerate(header_row, start=1):
                    header_value = cell.value
                    if isinstance(header_value, str) and header_value.lower() == "val":
                        val_column_index = col_index
                        break

                if val_column_index:
                    # Format values in the "Val" column as USD currency
                    for row in worksheet.iter_rows(min_row=start_row + 1):
                        cell = row[val_column_index - 1]
                        value = cell.value
                        if isinstance(value, (int, float)):
                            cell.number_format = numbers.FORMAT_CURRENCY_USD
                            cell.value = value

        # Save the modified workbook
        workbook.save(filename=file_path)

    def get_unique_sheet_name(self, workbook, base_name):
        """
        Generates a unique sheet name based on the provided base name.

        Args:
            workbook (xlsxwriter.Workbook): The workbook object.
            base_name (str): The base name for the sheet.

        Returns:
            str: The unique sheet name.

        Example:
            # Create an instance of the class
            obj = MyClass()

            # Generate a unique sheet name
            unique_name = obj.get_unique_sheet_name(workbook, "My Sheet")

            # Output: "My Sheet_1"
            print(unique_name)
        """
        worksheet_name = self.truncate_sheet_name(base_name)
        existing_sheet_names = [sheet.lower() for sheet in workbook.sheetnames]
        counter = 1
        while True:
            new_name = f"{worksheet_name}_{counter}"
            if new_name.lower() not in existing_sheet_names:
                worksheet_name = new_name
                break
            counter += 1
        return worksheet_name

    def create_hyperlinks(self,workbook, table_of_contents_worksheet):
        """
            Creates hyperlinks in the table of contents worksheet to navigate to other sheets in the workbook.

            Args:
                workbook (xlsxwriter.Workbook): The workbook object.
                table_of_contents_worksheet (xlsxwriter.Worksheet): The worksheet object for the table of contents.

            Example:
                # Create an instance of the class
                obj = MyClass()

                # Create hyperlinks in the table of contents worksheet
                obj.create_hyperlinks(workbook, table_of_contents_worksheet)
        """
        hyperlink_format = workbook.add_format({
            'font_color': 'blue',
            'underline': 1,
        })

        row_number = 1
        table_of_contents_worksheet.write('A1', 'No.')
        table_of_contents_worksheet.write('B1', 'Sheet Name')
        table_of_contents_worksheet.write('C1', 'Property Name')

        for worksheet_number, sheet_name in enumerate(workbook.sheetnames, start=1):
            if sheet_name != 'table_of_contents_worksheet':
                hyperlink_cell = f'B{row_number + 1}'
                table_of_contents_worksheet.write_url(
                    hyperlink_cell, f"internal:'{sheet_name}'!A1", string=sheet_name, cell_format=hyperlink_format)
                worksheet = workbook.get_worksheet_by_name(sheet_name)
                b1_value_formula = f'=\'{sheet_name}\'!B1'
                table_of_contents_worksheet.write(row_number, 2, b1_value_formula)
                table_of_contents_worksheet.write(row_number, 0, row_number)
                row_number += 1
                      
    def get_companyfacts_json_excel(self):
        """
            Retrieves company facts data in JSON format for each CIK number, generates an Excel workbook,
            and populates the worksheets with the data.

            Returns:
                None

            Example:
                # Create an instance of the class
                obj = CompanyFacts()

                # Set the company CIK numbers
                obj.company_CIKs = ["123456", "789012"]

                # Retrieve company facts data and generate Excel workbook
                obj.get_companyfacts_json_excel()
         """
        try:
            for company_cik in self.company_CIKs:
                json_data = self.retrieve_companyfacts_json(company_cik)
                if json_data is None:
                    continue
                
                Primary_Name = json_data['entityName']
                workbook_name = self.truncate_sheet_name(Primary_Name)               
                output_file_name = f"{workbook_name}.xlsx"
                sanitized_output_file_name = self.sanitize_filename(output_file_name)               
                output_file_path = os.path.join(self.folder_path, sanitized_output_file_name)               
                self.logger.info(f"Output File Path: {output_file_path}")
                
                # Check if the file exists and delete it if it does
                if os.path.isfile(output_file_path):
                    try:
                        os.remove(output_file_path)
                    except (PermissionError, IsADirectoryError):
                        self.logger.warning(f"Unable to delete file. Please close the file before running the script.")
                        continue
                
                # Create a new Workbook
                workbook = load_workbook()

                # Set option for NaN and infinity values to be treated as errors
                workbook.nan_inf_to_errors = True

                # Save the workbook to a file
                workbook.save(output_file_path)
                
                # Create a table of contents worksheet
                table_of_contents_worksheet = workbook.add_worksheet("table_of_contents_worksheet")
                
                # Iterate over the available JSON tree structures ('us-gaap', 'ifrs-full')
                for tree_name in ['us-gaap', 'ifrs-full']:
                    if tree_name in json_data['facts']:
                        tree_structure = tree_name
                        break
                else:
                    self.logger.warning(f"Unknown JSON tree structure for CIK: {company_cik}")
                    continue
                
                # Iterate over the items in the JSON tree structure
                for item in json_data['facts'][tree_structure]:
                    array_parameter_name = item
                    label_name = json_data['facts'][tree_structure][array_parameter_name]["label"]
                    description_name = json_data['facts'][tree_structure][array_parameter_name]['description']

                    for unit_item, unit_data in json_data['facts'][tree_structure][array_parameter_name]['units'].items():
                        row_index = 7
                        unique_worksheet_name = self.get_unique_sheet_name(workbook, array_parameter_name)
                        unit_worksheet = workbook.add_worksheet(unique_worksheet_name)
                        unit_worksheet.write(3, 1, unit_item)
                        unit_worksheet.write(0, 0, "array_parameter_name")
                        unit_worksheet.write(0, 1, array_parameter_name)
                        unit_worksheet.write(1, 0, "label_name")
                        unit_worksheet.write(1, 1, label_name)
                        unit_worksheet.write(2, 0, "description_name")
                        unit_worksheet.write(2, 1, description_name)
                        unit_worksheet.write(3, 0, "Unit")

                        df1 = pd.DataFrame(unit_data)

                        for col, value in enumerate(df1.columns):
                            unit_worksheet.write(row_index, col, value)
                        row_index += 1

                        for _, data_row in df1.iterrows():
                            for col, value in enumerate(data_row):
                                if pd.isna(value) or (isinstance(value, float) and not np.isfinite(value)):
                                    value = ""  # Set null values to an empty string
                                unit_worksheet.write(row_index, col, value)
                            row_index += 1

                        
                self.create_hyperlinks(workbook, table_of_contents_worksheet)
                # Save and close the workbook
                workbook.close()
                self.format_values_as_usd_excel(output_file_path)

        except Exception as e:
            self.logger.error(f"Error: {e}")
            sys.exit(1)
             
    def truncate_sheet_name(self, sheet_name):
        """
            Truncates the sheet name to a maximum of 31 characters and removes prohibited characters.

            Args:
                sheet_name (str): The original sheet name.

            Returns:
                str: The truncated and sanitized sheet name.

            Example:
                # Create an instance of the class
                obj = MyClass()

                # Truncate and sanitize the sheet name
                truncated_name = obj.truncate_sheet_name("My Sheet Name")

                # Output: "My Sheet Name"
                print(truncated_name)
        """
        sheet_name = re.sub(r'[<>:"/\\|?*]', '', sheet_name)
        return sheet_name[:26]

   # Get # Each entitys current filing history is available at the following URL:
    def Get_SubmissionContent_Json(self,db_path,Company_CIK_Number):

                    #Each entity�s current filing history is available at the following URL:
                    api_Submissions = f"https://data.sec.gov/submissions/CIK{Company_CIK_Number}.json"
                    #Sleep times of 10 seconds added in compliance with the SEC Regulations

                    try:
                        time.sleep(0.1)
                        response = requests.get(api_Submissions, headers={"User-Agent":self.user_agent})
                        response.raise_for_status()
                        json_object = json.loads(response.content)
                        time.sleep(0.1)
                    except requests.exceptions.RequestException as e:
                        logging.error(f"Error making API request: {str(e)}")
                        print(f'An error has occurred: {str(e)}\nLine number: {sys.exc_info()[-1].tb_lineno}')
   
                    #Opens the Webbrowser reflecting the data from which the extraction took place from 
                    time.sleep(0.1)
                    #webbrowser.open(api_Submissions)
                    time.sleep(0.1)
                    #prints the link to the data incase its requires
                
                    #below code uses the loads function (the s is used for strings, for objects or from files use load)
                    #the code loads the json data into a Python style dictionary
                    #Below variables are basically all the filing data that are provided in the link provided by the Sec
                    #not all are relevant but are kept for consistency
                    accessionNumber = []
                    filingDate= []
                    reportDate= []
                    acceptanceDateTime= []
                    act= []
                    form= []
                    fileNumber= []
                    filmNumber= []
                    items= []
                    size= []
                    isXBRL= []
                    isInlineXBRL= []
                    primaryDocument= []
                    primaryDocDescription= []
                    Document_Link = []
                    #the below basically goes through the Json Object extractts the data and stores them in arrays
                    #the code only goes up to a certain point, remaining data is stored in other JSON files that can be extracted
                    for item in json_object["filings"]["recent"]["accessionNumber"]:
                        accessionNumber.append(item)
                    for item in json_object["filings"]["recent"]["filingDate"]:
                        filingDate.append(item)
                    for item in json_object["filings"]["recent"]['reportDate']:
                        reportDate.append(item)
                    for item in json_object["filings"]["recent"]['acceptanceDateTime']:
                        acceptanceDateTime.append(item)
                    for item in json_object["filings"]["recent"]['act']:
                        act.append(item)
                    for item in json_object["filings"]["recent"]['form']:
                        form.append(item)
                    for item in json_object["filings"]["recent"]['fileNumber']:
                        fileNumber.append(item)
                    for item in json_object["filings"]["recent"]['filmNumber']:
                        filmNumber.append(item)
                    for item in json_object["filings"]["recent"]['items']:
                        items.append(item)
                    for item in json_object["filings"]["recent"]['size']:
                        size.append(item)
                    for item in json_object["filings"]["recent"]['isXBRL']:
                        isXBRL.append(item)
                    for item in json_object["filings"]["recent"]['isInlineXBRL']:
                        isInlineXBRL.append(item)
                    for item in json_object["filings"]["recent"]['primaryDocument']:
                        primaryDocument.append(item)
                    for item in json_object["filings"]["recent"]['primaryDocDescription']:
                        primaryDocDescription.append(item)
                    #gets the total number of rows to establish the end of a loop
                    number_values= len(accessionNumber)
                    #loop code below is used to create document links 
                    #the sec website provides the data in seperate files and to reach them the Acessision number
                    #excluding the special characters needs to be combined with the base-url and the document link to access the site
                    for i in range(0, (number_values)):
                        sec_base_url= 'https://www.sec.gov/Archives/edgar/data'
                        clean_accessionvalue = accessionNumber[i]
                        clean_primaryDocument = primaryDocument[i]
                        clean_accessionvalueReplaced = clean_accessionvalue.replace('-','')
                        clean_cik = Company_CIK_Number.replace("CIK",'')
                        Doc_Link = f'{sec_base_url}/{clean_cik}/{clean_accessionvalueReplaced}/{clean_primaryDocument}'
                        Document_Link.append(Doc_Link) 

                    #Script below uses Panda Library to read the values and store into a Date frame
                    df = pd.DataFrame({'accessionNumber': accessionNumber, 'filingDate': filingDate, 'reportDate': reportDate,
                                      'acceptanceDateTime': acceptanceDateTime, 'act': act,
                                      ' form':  form, 'fileNumber': fileNumber, 'filmNumber': filmNumber,
                                     'items': items, 'size':size, 'isXBRL': isXBRL, 'isInlineXBRL': isInlineXBRL,'primaryDocument':primaryDocument,
                                     'primaryDocDescription': primaryDocDescription, 'Document_Link': Document_Link})
                    # Opens a link to the datapath for the where the unique name was created
                    # From there it uses a panda stores dataframe to create  a table called filing list and replaces all the values if they exist
                    # then it closes the connection 
                    conn = sqlite3.connect(db_path)
                    df.to_sql("filing_list", conn, if_exists="replace")
                    conn.close()

class Filling_Links_Intial:
 
    def __init__(self, user_agent):
            self.user_agent = user_agent
            self.logger = logging.getLogger(__name__)
            
    def load_CIK_Values(self):
       
        time.sleep(0.1)
        api_url = f"https://www.sec.gov/files/company_tickers.json"
        print(api_url)
        try:
            response = requests.get(api_url, headers={"User-Agent":self.user_agent})
            response.raise_for_status()  # Raises an exception for non-200 status codes
            data = response.json()
            return data
        except requests.exceptions.RequestException as e:
            print(f"Warning: Error occurred while retrieving CIK Values")
            print(f"Error details: {str(e)}")
            return None
        except json.JSONDecodeError as e:
            print(f"Warning: Error occurred while parsing JSON response for CIK values")
            print(f"Error details: {str(e)}")
            return None